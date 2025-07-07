import pandas as pd
import requests
import time
import os
import unicodedata
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Arquivos
ARQUIVO_ENTRADA = 'planilhabase.xlsx'
ARQUIVO_SAIDA = 'planilha_corrigida.xlsx'
PAUSA = 0.5  # segundos

# Estilo de célula corrigida
fill_amarelo = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

def remover_acentos(texto):
    if isinstance(texto, str):
        return unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('utf-8')
    return texto

def consulta_cep(cep):
    cep = str(cep).replace('-', '').strip()
    if len(cep) != 8 or not cep.isdigit():
        return None
    url = f'https://viacep.com.br/ws/{cep}/json/'
    try:
        r = requests.get(url, timeout=2)
        if r.status_code == 200:
            data = r.json()
            if 'erro' not in data:
                return data
    except:
        return None
    return None

def buscar_cep_por_endereco(uf, cidade, logradouro):
    uf = uf.lower().strip()
    cidade = remover_acentos(cidade.lower().strip()).replace(" ", "+")
    logradouro = remover_acentos(logradouro.lower().strip()).replace(" ", "+")
    url = f"https://viacep.com.br/ws/{uf}/{cidade}/{logradouro}/json/"
    try:
        r = requests.get(url, timeout=5)
        if r.status_code == 200:
            data = r.json()
            if isinstance(data, list) and len(data) > 0:
                return data[0]
    except:
        return None
    return None

# Carrega planilha
df = pd.read_excel(ARQUIVO_ENTRADA)

# Garante coluna de status
if 'status' not in df.columns:
    df['status'] = ''

# Adiciona colunas de correção
colunas_corr = ['cep_corrigido', 'endereco_corrigido', 'bairro_corrigido', 'cidade_corrigida', 'uf_corrigido']
for col in colunas_corr:
    if col not in df.columns:
        df[col] = ''

print("🔁 Iniciando verificação de endereços...")

df_corrigido = df.copy()

for i, row in df_corrigido.iterrows():
    if str(row.get('status', '')).strip() != '':
        continue

    cep = str(row.get('CEP', '')).strip()
    print(f"[{i+1}/{len(df_corrigido)}] Consultando CEP {cep}... ", end='')

    resultado = consulta_cep(cep)
    time.sleep(PAUSA)

    if resultado:
        alterado = False

        endereco_planilha = str(row.get('Endereço', '')).lower()
        bairro_planilha = str(row.get('Bairro', '')).lower()
        cidade_planilha = str(row.get('Cidade', '')).lower()
        uf_planilha = str(row.get('UF', '')).upper()

        df_corrigido.at[i, 'cep_corrigido'] = resultado.get('cep', '').replace("-", "")

        if resultado.get('logradouro') and resultado['logradouro'].lower() not in endereco_planilha:
            df_corrigido.at[i, 'endereco_corrigido'] = resultado['logradouro']
            alterado = True
        else:
            df_corrigido.at[i, 'endereco_corrigido'] = row.get('Endereço', '')

        if resultado.get('bairro') and resultado['bairro'].lower() not in bairro_planilha:
            df_corrigido.at[i, 'bairro_corrigido'] = resultado['bairro']
            alterado = True
        else:
            df_corrigido.at[i, 'bairro_corrigido'] = row.get('Bairro', '')

        if resultado.get('localidade') and resultado['localidade'].lower() != cidade_planilha:
            df_corrigido.at[i, 'cidade_corrigida'] = resultado['localidade']
            alterado = True
        else:
            df_corrigido.at[i, 'cidade_corrigida'] = row.get('Cidade', '')

        if resultado.get('uf') and resultado['uf'].upper() != uf_planilha:
            df_corrigido.at[i, 'uf_corrigido'] = resultado['uf']
            alterado = True
        else:
            df_corrigido.at[i, 'uf_corrigido'] = row.get('UF', '')

        df_corrigido.at[i, 'status'] = 'CORRIGIDO' if alterado else 'OK'
        print("🔧 corrigido" if alterado else "✅ ok")

    else:
        print("❌ inválido, tentando por endereço... ", end='')
        uf = str(row.get('UF', '')).strip()
        cidade = str(row.get('Cidade', '')).strip()
        endereco = str(row.get('Endereço', '')).strip()

        resultado_endereco = buscar_cep_por_endereco(uf, cidade, endereco)
        time.sleep(PAUSA)

        if resultado_endereco:
            df_corrigido.at[i, 'cep_corrigido'] = resultado_endereco.get('cep', '').replace("-", "")
            df_corrigido.at[i, 'endereco_corrigido'] = resultado_endereco.get('logradouro', endereco)
            df_corrigido.at[i, 'bairro_corrigido'] = resultado_endereco.get('bairro', row.get('Bairro', ''))
            df_corrigido.at[i, 'cidade_corrigida'] = resultado_endereco.get('localidade', cidade)
            df_corrigido.at[i, 'uf_corrigido'] = resultado_endereco.get('uf', uf)
            df_corrigido.at[i, 'status'] = 'CORRIGIDO (endereço)'
            print("✅ corrigido via endereço")
        else:
            df_corrigido.at[i, 'status'] = 'CEP INVÁLIDO'
            print("❌ falha total")

    # Salvamento a cada 10 linhas
    if i % 10 == 0:
        df_corrigido.to_excel(ARQUIVO_SAIDA, index=False)

# Padronização: remover acentos e deixar maiúsculas
for col in df_corrigido.columns:
    if df_corrigido[col].dtype == object:
        df_corrigido[col] = df_corrigido[col].apply(remover_acentos).str.upper()

# Remove traço do CEP corrigido
df_corrigido['cep_corrigido'] = df_corrigido['cep_corrigido'].astype(str).str.replace("-", "", regex=False)

# Salva versão final
df_corrigido.to_excel(ARQUIVO_SAIDA, index=False)

# Backup com timestamp
timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
backup_file = f"planilha_corrigida_{timestamp}.xlsx"
df_corrigido.to_excel(backup_file, index=False)
print(f"📁 Backup salvo como: {backup_file}")

# Formatação visual no Excel
wb = load_workbook(ARQUIVO_SAIDA)
ws = wb.active
colunas = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}

for row in range(2, ws.max_row + 1):
    if ws.cell(row, colunas['status']).value in ['CORRIGIDO', 'CORRIGIDO (ENDERECO)']:
        for campo in colunas_corr:
            if campo in colunas:
                ws.cell(row, colunas[campo]).fill = fill_amarelo

wb.save(ARQUIVO_SAIDA)
print("✅ Arquivo final salvo e formatado com destaques.")

